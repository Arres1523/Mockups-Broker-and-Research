#!/usr/bin/env python3
"""Load executive summary rows from CSV/XLSX into Supabase."""

from __future__ import annotations

import argparse
import csv
import json
import os
import sys
import urllib.error
import urllib.request
from pathlib import Path
from typing import Iterable

EXPECTED_COLUMNS = [
    "property_name",
    "market",
    "broker_name",
    "units",
    "year_built",
    "physical_occupancy_pct",
    "noi_actual",
    "expense_ratio_actual",
    "dscr_current",
    "in_place_annual_upside",
]

NUMERIC_COLUMNS = {
    "units": int,
    "year_built": int,
    "physical_occupancy_pct": float,
    "noi_actual": float,
    "expense_ratio_actual": float,
    "dscr_current": float,
    "in_place_annual_upside": float,
}


def normalize_number(raw_value: str | None, caster):
    if raw_value is None:
        return None

    value = str(raw_value).strip()
    if not value:
        return None

    cleaned = value.replace("$", "").replace(",", "").replace("%", "").replace(" ", "")
    return caster(cleaned)


def normalize_row(row: dict[str, object]) -> dict[str, object]:
    normalized = {}

    for column in EXPECTED_COLUMNS:
        value = row.get(column)

        if column in NUMERIC_COLUMNS:
            normalized[column] = normalize_number(value, NUMERIC_COLUMNS[column])
        else:
            normalized[column] = str(value).strip() if value not in (None, "") else None

    if not normalized["property_name"]:
        raise ValueError("property_name is required for every row")

    return normalized


def read_csv_rows(path: Path) -> list[dict[str, object]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        missing = [column for column in EXPECTED_COLUMNS if column not in reader.fieldnames]
        if missing:
            raise ValueError(f"Missing columns in CSV: {', '.join(missing)}")

        return [normalize_row(row) for row in reader]


def read_xlsx_rows(path: Path) -> list[dict[str, object]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Excel support requires openpyxl. Install it with: pip install openpyxl"
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    missing = [column for column in EXPECTED_COLUMNS if column not in headers]
    if missing:
        raise ValueError(f"Missing columns in XLSX: {', '.join(missing)}")

    rows = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        rows.append(normalize_row(dict(zip(headers, values))))
    return rows


def read_rows(path: Path) -> list[dict[str, object]]:
    if path.suffix.lower() == ".csv":
        return read_csv_rows(path)
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return read_xlsx_rows(path)
    raise ValueError("Unsupported file type. Use .csv, .xlsx, or .xlsm")


def chunked(rows: list[dict[str, object]], size: int) -> Iterable[list[dict[str, object]]]:
    for start in range(0, len(rows), size):
        yield rows[start : start + size]


def upload_rows(rows: list[dict[str, object]], chunk_size: int) -> None:
    supabase_url = os.environ.get("SUPABASE_URL")
    service_role_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")

    if not supabase_url or not service_role_key:
        raise RuntimeError(
            "SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY must be set for API upload."
        )

    endpoint = f"{supabase_url.rstrip('/')}/rest/v1/executive_summaries"

    for payload in chunked(rows, chunk_size):
        request = urllib.request.Request(
            endpoint,
            data=json.dumps(payload).encode("utf-8"),
            headers={
                "Content-Type": "application/json",
                "apikey": service_role_key,
                "Authorization": f"Bearer {service_role_key}",
                "Prefer": "resolution=merge-duplicates,return=minimal",
            },
            method="POST",
        )

        with urllib.request.urlopen(request) as response:
            if response.status not in {200, 201, 204}:
                raise RuntimeError(f"Unexpected Supabase response: {response.status}")


def write_sql(rows: list[dict[str, object]], output_path: Path) -> None:
    values_sql = []
    for row in rows:
        values = []
        for column in EXPECTED_COLUMNS:
            value = row[column]
            if value is None:
                values.append("null")
            elif isinstance(value, str):
                values.append("'" + value.replace("'", "''") + "'")
            else:
                values.append(str(value))
        values_sql.append(f"({', '.join(values)})")

    statement = (
        "insert into public.executive_summaries (\n  "
        + ",\n  ".join(EXPECTED_COLUMNS)
        + "\n)\nvalues\n  "
        + ",\n  ".join(values_sql)
        + "\n;"
    )
    output_path.write_text(statement, encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input_file", type=Path, help="CSV or XLSX with executive summary rows")
    parser.add_argument(
        "--mode",
        choices=["api", "sql"],
        default="api",
        help="Upload to Supabase over the REST API or generate a SQL seed file",
    )
    parser.add_argument(
        "--chunk-size",
        type=int,
        default=250,
        help="Rows per request when using API mode",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("executive_summaries_seed.sql"),
        help="SQL output path when --mode sql is selected",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    rows = read_rows(args.input_file)

    if args.mode == "sql":
        write_sql(rows, args.output)
        print(f"Wrote {len(rows)} rows to {args.output}")
        return 0

    upload_rows(rows, args.chunk_size)
    print(f"Uploaded {len(rows)} rows to Supabase")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (RuntimeError, ValueError, urllib.error.URLError) as exc:
        print(f"Error: {exc}", file=sys.stderr)
        raise SystemExit(1)
